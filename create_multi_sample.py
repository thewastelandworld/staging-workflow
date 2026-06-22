import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()

C_HDR  = "2563EB"
C_DESC = "FEF9C3"
C_ALT  = "EFF6FF"

def thin():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def desc_cell(ws, row, col, val, bg=C_DESC, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, name="Noto Sans JP", size=10,
                  color="92400E" if bg == C_DESC else "111827")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = thin()

def hdr_cell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=C_HDR)
    c.font = Font(bold=True, color="FFFFFF", name="Noto Sans JP", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()

def data_cell(ws, row, col, val, alt=False):
    c = ws.cell(row=row, column=col, value=val)
    if alt:
        c.fill = PatternFill("solid", fgColor=C_ALT)
    c.font = Font(name="Noto Sans JP", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = thin()

# ステージ名 | 説明 | 担当チーム | 確認チーム | 確認内容 | 期限
HEADERS = ["ステージ名", "説明", "担当チーム", "確認チーム", "確認内容", "期限"]
WIDTHS  = [28, 38, 20, 20, 36, 14]

def make_sheet(ws, case_description, stages):
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # Row 1: case description (merged B1:F1)
    ws.row_dimensions[1].height = 36
    desc_cell(ws, 1, 1, "ケース説明", bold=True)
    ws.merge_cells("B1:F1")
    desc_cell(ws, 1, 2, case_description)
    for col in range(3, 7):
        c = ws.cell(row=1, column=col)
        c.border = thin()
        c.fill = PatternFill("solid", fgColor=C_DESC)

    # Row 2: headers
    ws.row_dimensions[2].height = 22
    for col, h in enumerate(HEADERS, start=1):
        hdr_cell(ws, 2, col, h)

    # Row 3+: stage data (name, desc, team, reviewer_team, check_content, deadline)
    for ri, stage in enumerate(stages, start=3):
        ws.row_dimensions[ri].height = 38
        for ci, val in enumerate(stage, start=1):
            data_cell(ws, ri, ci, val, alt=(ri % 2 == 0))


# ── Sheet 1: 新規会員登録フロー ──────────────────────────────
ws1 = wb.active
ws1.title = "新規会員登録フロー"
make_sheet(ws1,
    "新規ユーザーがメールアドレスで会員登録し、本人確認・審査を経てアカウントを開設するフロー。",
    [
        ("メールアドレス確認",
         "入力フォームのバリデーション。重複チェック・形式確認。",
         "フロントエンドチーム", "QAチーム",
         "入力値が正しい形式であること。重複ユーザーが存在しないこと。",
         "2026-07-05"),
        ("会員情報入力",
         "氏名・住所・生年月日等の個人情報入力フォーム表示。",
         "フロントエンドチーム", "QAチーム",
         "必須項目がすべて入力されていること。住所の郵便番号自動補完が動作すること。",
         "2026-07-08"),
        ("本人確認書類アップロード",
         "身分証明書（運転免許証・マイナンバー等）のアップロード・OCR読取。",
         "バックエンドチーム", "セキュリティチーム",
         "書類画像が鮮明であること。OCR読取結果と入力情報が一致すること。有効期限内であること。",
         "2026-07-10"),
        ("審査・承認処理",
         "管理者による申請内容の最終確認・承認。不備がある場合は差し戻し。",
         "オペレーションチーム", "経理チーム",
         "申請内容に虚偽がないこと。反社チェックが完了していること。",
         "2026-07-12"),
        ("アカウント開設完了通知",
         "承認後にウェルカムメール送信。マイページへのログイン案内を含む。",
         "バックエンドチーム", "QAチーム",
         "メールが正常に届くこと。ログインURLが正しいこと。初期パスワードが設定されていること。",
         "2026-07-15"),
    ]
)

# ── Sheet 2: 商品返品・返金フロー ────────────────────────────
ws2 = wb.create_sheet("商品返品・返金フロー")
make_sheet(ws2,
    "購入者からの返品申請を受け付け、商品検品・返金額計算・クレジット返金処理を行うフロー。",
    [
        ("返品申請受付",
         "返品理由・商品状態・注文番号をWebフォームで受付。",
         "フロントエンドチーム", "QAチーム",
         "申請フォームが正常に送信されること。注文番号の存在確認が行われること。",
         "2026-07-06"),
        ("返品商品の検品",
         "届いた返品商品の状態確認・写真記録。不良品か使用済みかを判定。",
         "オペレーションチーム", "経理チーム",
         "商品状態が記録されていること。写真が添付されていること。判定結果が入力されていること。",
         "2026-07-09"),
        ("返金額の計算",
         "商品価格から送料・手数料を控除した返金額を算出。",
         "経理チーム", "経理チーム",
         "返金額の計算式が正しいこと。控除項目が明細に記載されていること。",
         "2026-07-11"),
        ("クレジット返金処理",
         "カード会社への返金申請。承認後3〜5営業日で反映。",
         "決済チーム", "経理チーム",
         "返金申請がカード会社に送信されていること。取引IDが記録されていること。",
         "2026-07-14"),
        ("返金完了通知",
         "返金完了メールをユーザーへ送信。返金明細・完了日を記載。",
         "バックエンドチーム", "QAチーム",
         "メールに返金額・返金日・明細が含まれていること。",
         "2026-07-16"),
    ]
)

# ── Sheet 3: ポイント交換フロー ──────────────────────────────
ws3 = wb.create_sheet("ポイント交換フロー")
make_sheet(ws3,
    "保有ポイントを使って交換カタログから商品を選び、発送手配・完了通知を行うフロー。",
    [
        ("ポイント残高確認",
         "交換可能ポイント数・有効期限を取得して画面表示。",
         "バックエンドチーム", "QAチーム",
         "残高が正確に表示されること。有効期限が近いポイントが優先表示されること。",
         "2026-07-07"),
        ("交換商品の選択",
         "交換カタログから商品選択。必要ポイント数と在庫をリアルタイム確認。",
         "フロントエンドチーム", "QAチーム",
         "在庫切れ商品が選択不可になっていること。必要ポイント数が正しく表示されること。",
         "2026-07-10"),
        ("交換申請・ポイント引落",
         "選択商品のポイントをアカウントから即時引き落とし。",
         "バックエンドチーム", "経理チーム",
         "ポイントが正確に引き落とされていること。残高が更新されていること。二重引き落としがないこと。",
         "2026-07-12"),
        ("発送手配",
         "倉庫システムへ発送指示を連携。送り状番号を取得・登録。",
         "オペレーションチーム", "QAチーム",
         "倉庫システムへの連携が完了していること。送り状番号が登録されていること。",
         "2026-07-15"),
        ("発送完了通知",
         "追跡番号付きの発送完了メールをユーザーへ送信。",
         "バックエンドチーム", "QAチーム",
         "追跡URLが有効であること。メール内の商品名・配送先が正しいこと。",
         "2026-07-18"),
    ]
)

out = "sample_multi_sheet_import.xlsx"
wb.save(out)
print(f"Saved: {out}")
